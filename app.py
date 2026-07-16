import os
import sys
import json
import openpyxl
import webbrowser
from threading import Timer
from flask import Flask, jsonify, request, send_from_directory, session
from werkzeug.utils import secure_filename

from docx_generator import generate_docx_labels
from pdf_generator import generate_printing_slip, generate_commercial_invoice

app = Flask(__name__, static_folder='static', static_url_path='')
app.secret_key = os.environ.get('SECRET_KEY', 'rms-warehouse-secret-key-123')

# User database configuration
USERS = {
    'raj': {'password': 'Plainfield1$', 'role': 'admin'},
    'sales': {'password': 'Rms12361$', 'role': 'admin'},
    'accounting': {'password': 'Acc12361$', 'role': 'employee'},
    'warehouse': {'password': 'Rhea12361$', 'role': 'employee'}
}

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Global Database States (Thread-safe-ish for local single-user app)
CURRENT_DB_NAME = None
CURRENT_DB_PATH = None
CURRENT_SHEET = None
EXCEL_SHEETS = []
CURRENT_HEADER_ROW = 1

CONFIG_PATH = os.path.join(BASE_DIR, "db_config.json")

def load_db_config():
    global CURRENT_DB_NAME, CURRENT_DB_PATH, CURRENT_SHEET, EXCEL_SHEETS, CURRENT_HEADER_ROW
    if os.path.exists(CONFIG_PATH):
        try:
            with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
                config = json.load(f)
            
            filename = config.get('filename')
            if filename:
                file_path = os.path.join(BASE_DIR, filename)
                if os.path.exists(file_path):
                    CURRENT_DB_NAME = filename
                    CURRENT_DB_PATH = file_path
                    CURRENT_SHEET = config.get('sheet_name')
                    EXCEL_SHEETS = config.get('sheets', [])
                    CURRENT_HEADER_ROW = config.get('header_row', 1)
                    print(f"Loaded database configuration from config file: {filename} (Sheet: {CURRENT_SHEET}, Header: {CURRENT_HEADER_ROW})")
                    return
        except Exception as e:
            print(f"Error loading db_config.json: {e}")
            
    # Default fallback if config load fails or config doesn't exist
    CURRENT_DB_NAME = "Warehouse_Tracking_sheet_1.xlsx"
    CURRENT_DB_PATH = os.path.join(BASE_DIR, CURRENT_DB_NAME)
    CURRENT_SHEET = "Main"
    EXCEL_SHEETS = ["Main"]
    CURRENT_HEADER_ROW = 1


def save_db_config():
    try:
        config = {
            'filename': CURRENT_DB_NAME,
            'sheet_name': CURRENT_SHEET,
            'sheets': EXCEL_SHEETS,
            'header_row': CURRENT_HEADER_ROW
        }
        with open(CONFIG_PATH, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=4)
        print(f"Saved database configuration to config file: {CURRENT_DB_NAME}")
    except Exception as e:
        print(f"Error saving db_config.json: {e}")

# Load configuration on startup
load_db_config()

TEMPLATE_DOCX_PATH = os.path.join(BASE_DIR, "GE Label - 42300276510.docx")
OUTPUTS_DIR = os.path.join(BASE_DIR, "Outputs")

# Ensure outputs directory exists
os.makedirs(OUTPUTS_DIR, exist_ok=True)


def calculate_record_hash(rec):
    keys = [
        'inbound_date', 'rms_po', 'part_received', 'vendor', 'promise_date',
        'inbound_notes', 'vendor_contact', 'received_date', 'inbound_carrier', 'inbound_tracking',
        'inbound_l', 'inbound_w', 'inbound_h', 'inbound_weight', 'inbound_charges',
        'outbound_date', 'customer', 'customer_po', 'rms_invoice', 'ship_to',
        'line_num', 'hs_code', 'shipped_date', 'invoice_status', 'outbound_l',
        'outbound_w', 'outbound_h', 'outbound_weight', 'outbound_carrier', 'outbound_tracking',
        'crating_charges', 'shipping_charges', 'customer_contact', 'outbound_notes',
        'no_of_boxes', 'photo', 'report'
    ]
    content = "|".join(str(rec.get(k, "")).strip() for k in keys)
    import hashlib
    return hashlib.md5(content.encode('utf-8')).hexdigest()


def get_excel_row_record(sheet, row_idx, boxes_col_idx, photo_col_idx, report_col_idx):
    row = []
    for c in range(1, 35):
        val = sheet.cell(row=row_idx, column=c).value
        if val is None:
            row.append("")
        elif isinstance(val, float) and val.is_integer():
            row.append(str(int(val)))
        else:
            val_str = str(val)
            if val_str.endswith(" 00:00:00"):
                val_str = val_str[:-9]
            row.append(val_str)
            
    if len(row) < 34:
        row = row + [""] * (34 - len(row))
        
    no_of_boxes = ""
    if boxes_col_idx != -1 and boxes_col_idx <= sheet.max_column:
        val = sheet.cell(row=row_idx, column=boxes_col_idx).value
        no_of_boxes = str(val).strip() if val is not None else ""
        
    photo = ""
    if photo_col_idx != -1 and photo_col_idx <= sheet.max_column:
        val = sheet.cell(row=row_idx, column=photo_col_idx).value
        photo = str(val).strip() if val is not None else ""
        
    report = ""
    if report_col_idx != -1 and report_col_idx <= sheet.max_column:
        val = sheet.cell(row=row_idx, column=report_col_idx).value
        report = str(val).strip() if val is not None else ""
        
    rec = {
        'inbound_date': row[0].strip(),
        'rms_po': row[1].strip(),
        'part_received': row[2].strip(),
        'vendor': row[3].strip(),
        'promise_date': row[4].strip(),
        'inbound_notes': row[5].strip(),
        'vendor_contact': row[6].strip(),
        'received_date': row[7].strip(),
        'inbound_carrier': row[8].strip(),
        'inbound_tracking': row[9].strip(),
        'inbound_l': row[10].strip(),
        'inbound_w': row[11].strip(),
        'inbound_h': row[12].strip(),
        'inbound_weight': row[13].strip(),
        'inbound_charges': row[14].strip(),
        'outbound_date': row[15].strip().replace(" 00:00:00", ""),
        'customer': row[16].strip(),
        'customer_po': row[17].strip(),
        'rms_invoice': row[18].strip(),
        'ship_to': row[19].strip(),
        'line_num': row[20].strip(),
        'hs_code': row[21].strip(),
        'shipped_date': row[22].strip(),
        'invoice_status': row[23].strip(),
        'outbound_l': row[24].strip(),
        'outbound_w': row[25].strip(),
        'outbound_h': row[26].strip(),
        'outbound_weight': row[27].strip(),
        'outbound_carrier': row[28].strip(),
        'outbound_tracking': row[29].strip(),
        'crating_charges': row[30].strip(),
        'shipping_charges': row[31].strip(),
        'customer_contact': row[32].strip(),
        'outbound_notes': row[33].strip(),
        'no_of_boxes': no_of_boxes,
        'photo': photo,
        'report': report,
    }
    return rec


def parse_excel_database(file_path, sheet_name=None, header_row=1):
    """
    Parses the warehouse tracking Excel sheet.
    Maps columns by absolute index to match the CSV format.
    """
    records = []
    if not os.path.exists(file_path):
        print(f"Warning: Excel file not found at {file_path}")
        return records
        
    wb = None
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        if not sheet_name or sheet_name not in wb.sheetnames:
            sheet = wb.active
        else:
            sheet = wb[sheet_name]
            
        row_iterator = sheet.iter_rows(values_only=True)
        
        # Ensure header_row is within valid bounds
        h_idx = max(1, int(header_row))
        
        headers = None
        for _ in range(h_idx):
            try:
                headers = next(row_iterator)
            except StopIteration:
                break
                
        if not headers:
            return records
            
        headers_clean = [str(h).strip().lower() if h is not None else "" for h in headers]
        boxes_idx = next((i for i, h in enumerate(headers_clean) if "boxes" in h or "no. of boxes" in h), -1)
        photo_idx = next((i for i, h in enumerate(headers_clean) if "photo" in h), -1)
        report_idx = next((i for i, h in enumerate(headers_clean) if "report" in h), -1)
        
        consecutive_empty_count = 0
        
        # Iterate data rows starting after the header
        for idx, row_cells in enumerate(row_iterator):
            row = []
            for val in row_cells:
                if val is None:
                    row.append("")
                elif isinstance(val, float) and val.is_integer():
                    # Format float representing integers as integers (e.g. 9141.0 -> "9141")
                    row.append(str(int(val)))
                else:
                    # Convert cell value to clean string
                    val_str = str(val)
                    if val_str.endswith(" 00:00:00"):
                        val_str = val_str[:-9]
                    row.append(val_str)
                    
            if not row or all(cell.strip() == "" for cell in row):
                consecutive_empty_count += 1
                if consecutive_empty_count >= 100:
                    break
                continue
                
            consecutive_empty_count = 0  # Reset counter for non-empty row
            
            # Skip rows containing "cancelled" (case-insensitive)
            if any("cancelled" in cell.lower() for cell in row):
                continue
                
            max_len = max(34, boxes_idx + 1, photo_idx + 1, report_idx + 1)
            if len(row) < max_len:
                row = row + [""] * (max_len - len(row))
                
            record = {
                'row_id': idx + h_idx + 1, # Line number in sheet (1-indexed, starts at header_row + 1)
                'inbound_date': row[0].strip(),
                'rms_po': row[1].strip(),
                'part_received': row[2].strip(),
                'vendor': row[3].strip(),
                'promise_date': row[4].strip(),
                'inbound_notes': row[5].strip(),
                'vendor_contact': row[6].strip(),
                'received_date': row[7].strip(),
                'inbound_carrier': row[8].strip(),
                'inbound_tracking': row[9].strip(),
                'inbound_l': row[10].strip(),
                'inbound_w': row[11].strip(),
                'inbound_h': row[12].strip(),
                'inbound_weight': row[13].strip(),
                'inbound_charges': row[14].strip(),
                'outbound_date': row[15].strip().replace(" 00:00:00", ""),
                'customer': row[16].strip(),
                'customer_po': row[17].strip(),
                'rms_invoice': row[18].strip(),
                'ship_to': row[19].strip(),
                'line_num': row[20].strip(),
                'hs_code': row[21].strip(),
                'shipped_date': row[22].strip(),
                'invoice_status': row[23].strip(),
                'outbound_l': row[24].strip(),
                'outbound_w': row[25].strip(),
                'outbound_h': row[26].strip(),
                'outbound_weight': row[27].strip(),
                'outbound_carrier': row[28].strip(),
                'outbound_tracking': row[29].strip(),
                'crating_charges': row[30].strip(),
                'shipping_charges': row[31].strip(),
                'customer_contact': row[32].strip(),
                'outbound_notes': row[33].strip(),
                'no_of_boxes': row[boxes_idx].strip() if boxes_idx != -1 else "",
                'photo': row[photo_idx].strip() if photo_idx != -1 else "",
                'report': row[report_idx].strip() if report_idx != -1 else "",
            }
            record['row_hash'] = calculate_record_hash(record)
            
            if record['rms_po'] or record['customer_po'] or record['part_received']:
                records.append(record)
                
    except Exception as e:
        print(f"Error parsing Excel: {e}")
    finally:
        if wb:
            wb.close()
            
    records.reverse()
    return records


def load_current_database():
    """Loads records from the active database file."""
    if not CURRENT_DB_PATH or not os.path.exists(CURRENT_DB_PATH):
        print(f"Warning: Database path does not exist: {CURRENT_DB_PATH}")
        return []
        
    ext = os.path.splitext(CURRENT_DB_PATH)[1].lower()
    if ext in ['.xlsx', '.xls']:
        return parse_excel_database(CURRENT_DB_PATH, CURRENT_SHEET, CURRENT_HEADER_ROW)
    return []


@app.route('/api/login', methods=['POST'])
def login():
    data = request.json or {}
    username = data.get('username')
    password = data.get('password')
    
    if not username or not password:
        return jsonify({'error': 'Username and password are required.'}), 400
        
    user = USERS.get(username)
    if user and user['password'] == password:
        session['username'] = username
        session['role'] = user['role']
        return jsonify({
            'success': True,
            'username': username,
            'role': user['role']
        })
        
    return jsonify({'error': 'Invalid username or password.'}), 401


@app.route('/api/logout', methods=['POST', 'GET'])
def logout():
    session.clear()
    return jsonify({'success': True})


@app.route('/api/auth-status', methods=['GET'])
def auth_status():
    if 'username' in session:
        return jsonify({
            'authenticated': True,
            'username': session['username'],
            'role': session.get('role', 'employee')
        })
    return jsonify({'authenticated': False})


@app.before_request
def check_auth():
    if app.testing:
        return
        
    # Exclude auth APIs
    if request.path.startswith('/api/') and request.path not in ['/api/login', '/api/logout', '/api/auth-status']:
        username = session.get('username')
        if not username:
            return jsonify({'error': 'Unauthorized. Please log in.'}), 401
            
        # Role-based restriction:
        # "admin: raj and sales, can access everything or all three section i.e. Shipping, Receiving and Order Entry"
        # "employee: accounting and warehousing, can only access Shipping and Receiving section"
        if request.path == '/api/save-order':
            role = session.get('role')
            if role != 'admin':
                return jsonify({'error': 'Forbidden. Only administrators can access this section.'}), 403


@app.route('/')
def index():
    return send_from_directory('static', 'index.html')


@app.after_request
def add_header(response):
    """Disable caching for development"""
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '-1'
    return response


@app.route('/api/records')
def get_records():
    records = load_current_database()
    return jsonify(records)


@app.route('/api/db-status')
def get_db_status():
    return jsonify({
        'filename': CURRENT_DB_NAME,
        'sheet_name': CURRENT_SHEET,
        'sheets': EXCEL_SHEETS,
        'header_row': CURRENT_HEADER_ROW
    })


@app.route('/api/pick-file', methods=['POST', 'GET'])
def pick_file():
    global CURRENT_DB_PATH, CURRENT_DB_NAME, CURRENT_SHEET, EXCEL_SHEETS, CURRENT_HEADER_ROW
    import platform
    import subprocess
    
    file_path = None
    system_os = platform.system()
    
    if system_os == 'Darwin':
        # macOS: Run AppleScript to open a native file dialog
        script = 'POSIX path of (choose file of type {"xlsx", "xls", "org.openxmlformats.spreadsheetml.sheet", "com.microsoft.excel.xls"} with prompt "Select Excel Database File")'
        try:
            proc = subprocess.run(['osascript', '-e', script], capture_output=True, text=True)
            if proc.returncode == 0:
                file_path = proc.stdout.strip()
        except Exception as e:
            print(f"Error opening macOS dialog: {e}")
            
    elif system_os == 'Windows':
        # Windows: Run PowerShell System.Windows.Forms dialog
        ps_script = (
            "[System.Reflection.Assembly]::LoadWithPartialName('System.Windows.Forms') | Out-Null;"
            "$dialog = New-Object System.Windows.Forms.OpenFileDialog;"
            "$dialog.Filter = 'Excel Files (*.xlsx; *.xls)|*.xlsx;*.xls';"
            "$dialog.Title = 'Select Excel Database File';"
            "$res = $dialog.ShowDialog();"
            "if ($res -eq 'OK') { Write-Output $dialog.FileName }"
        )
        try:
            proc = subprocess.run(
                ['powershell', '-NoProfile', '-Command', ps_script],
                capture_output=True,
                text=True,
                creationflags=0x08000000 # CREATE_NO_WINDOW
            )
            if proc.returncode == 0:
                file_path = proc.stdout.strip()
        except Exception as e:
            print(f"Error opening Windows dialog: {e}")
            
    if not file_path:
        return jsonify({'cancelled': True})
        
    if not os.path.exists(file_path):
        return jsonify({'error': 'Selected file does not exist.'}), 400
        
    filename = os.path.basename(file_path)
    ext = os.path.splitext(filename)[1].lower()
    
    if ext not in ['.xlsx', '.xls']:
        return jsonify({'error': 'Only Excel files (.xlsx, .xls) are supported.'}), 400
        
    CURRENT_DB_NAME = filename
    CURRENT_DB_PATH = file_path
    CURRENT_SHEET = None
    EXCEL_SHEETS = []
    CURRENT_HEADER_ROW = 1
    
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        
        EXCEL_SHEETS = sheets
        if len(sheets) > 0:
            CURRENT_SHEET = sheets[0]
            
        save_db_config()
        return jsonify({
            'success': True,
            'file_path': file_path,
            'filename': filename,
            'sheets': sheets,
            'selected_sheet': CURRENT_SHEET,
            'header_row': CURRENT_HEADER_ROW
        })
    except Exception as e:
        return jsonify({'error': f"Failed to read Excel file: {str(e)}"}), 500


@app.route('/api/upload-database', methods=['POST'])
def upload_database():
    global CURRENT_DB_PATH, CURRENT_DB_NAME, CURRENT_SHEET, EXCEL_SHEETS, CURRENT_HEADER_ROW
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
        
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
        
    if file:
        filename = secure_filename(file.filename)
        file_path = os.path.join(BASE_DIR, filename)
        file.save(file_path)
        
        ext = os.path.splitext(filename)[1].lower()
        
        CURRENT_DB_NAME = filename
        CURRENT_DB_PATH = file_path
        CURRENT_SHEET = None
        EXCEL_SHEETS = []
        CURRENT_HEADER_ROW = 1
        
        if ext in ['.xlsx', '.xls']:
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True)
                sheets = wb.sheetnames
                wb.close()
                
                EXCEL_SHEETS = sheets
                if len(sheets) > 0:
                    CURRENT_SHEET = sheets[0]
                    records = parse_excel_database(file_path, CURRENT_SHEET, CURRENT_HEADER_ROW)
                else:
                    records = []
                    
                save_db_config()
                return jsonify({
                    'success': True,
                    'type': 'excel',
                    'filename': filename,
                    'sheets': sheets,
                    'selected_sheet': CURRENT_SHEET,
                    'header_row': CURRENT_HEADER_ROW,
                    'records': records
                })
            except Exception as e:
                return jsonify({'error': f"Failed to read Excel workbook: {str(e)}"}), 500
        else:
            return jsonify({'error': 'Unsupported file format. Please upload Excel (.xlsx, .xls) files.'}), 400


@app.route('/api/load-sheet', methods=['POST'])
def load_sheet():
    global CURRENT_SHEET, CURRENT_HEADER_ROW
    data = request.json
    if not data:
        return jsonify({'error': 'No data provided'}), 400
        
    # Switch sheet if provided
    sheet_name = data.get('sheet_name', CURRENT_SHEET)
    if sheet_name not in EXCEL_SHEETS:
        return jsonify({'error': f"Sheet '{sheet_name}' not found in workbook."}), 400
        
    # Read header row if provided
    header_row_val = data.get('header_row', CURRENT_HEADER_ROW)
    try:
        CURRENT_HEADER_ROW = max(1, int(header_row_val))
    except (ValueError, TypeError):
        pass
        
    CURRENT_SHEET = sheet_name
    records = parse_excel_database(CURRENT_DB_PATH, CURRENT_SHEET, CURRENT_HEADER_ROW)
    save_db_config()
    return jsonify({
        'success': True,
        'filename': CURRENT_DB_NAME,
        'sheet_name': CURRENT_SHEET,
        'header_row': CURRENT_HEADER_ROW,
        'records': records
    })


@app.route('/api/save-order', methods=['POST'])
def save_order():
    global CURRENT_DB_PATH, CURRENT_SHEET, CURRENT_HEADER_ROW
    data = request.json
    if not data:
        return jsonify({'error': 'No data provided'}), 400
        
    if not CURRENT_DB_PATH or not os.path.exists(CURRENT_DB_PATH):
        return jsonify({'error': 'No active database file loaded.'}), 400
        
    # Ensure 34 columns in the exact order as defined
    row_data = [
        data.get('inbound_date', '').strip(),
        data.get('rms_po', '').strip(),
        data.get('part_received', '').strip(),
        data.get('vendor', '').strip(),
        data.get('promise_date', '').strip(),
        data.get('inbound_notes', '').strip(),
        data.get('vendor_contact', '').strip(),
        data.get('received_date', '').strip(),
        data.get('inbound_carrier', '').strip(),
        data.get('inbound_tracking', '').strip(),
        data.get('inbound_l', '').strip(),
        data.get('inbound_w', '').strip(),
        data.get('inbound_h', '').strip(),
        data.get('inbound_weight', '').strip(),
        data.get('inbound_charges', '').strip(),
        data.get('outbound_date', '').strip(),
        data.get('customer', '').strip(),
        data.get('customer_po', '').strip(),
        data.get('rms_invoice', '').strip(),
        data.get('ship_to', '').strip(),
        data.get('line_num', '').strip(),
        data.get('hs_code', '').strip(),
        data.get('shipped_date', '').strip(),
        data.get('invoice_status', '').strip(),
        data.get('outbound_l', '').strip(),
        data.get('outbound_w', '').strip(),
        data.get('outbound_h', '').strip(),
        data.get('outbound_weight', '').strip(),
        data.get('outbound_carrier', '').strip(),
        data.get('outbound_tracking', '').strip(),
        data.get('crating_charges', '').strip(),
        data.get('shipping_charges', '').strip(),
        data.get('customer_contact', '').strip(),
        data.get('outbound_notes', '').strip(),
    ]
    
    ext = os.path.splitext(CURRENT_DB_PATH)[1].lower()
    try:
        if ext in ['.xlsx', '.xls']:
            wb = openpyxl.load_workbook(CURRENT_DB_PATH)
            if not CURRENT_SHEET or CURRENT_SHEET not in wb.sheetnames:
                sheet = wb.active
            else:
                sheet = wb[CURRENT_SHEET]
                
            # Helper to clean/cast values for Excel
            def cast_val(val):
                val_str = str(val).strip()
                if val_str == "":
                    return None
                try:
                    if val_str.isdigit():
                        return int(val_str)
                    float_val = float(val_str)
                    if float_val.is_integer():
                        return int(float_val)
                    return float_val
                except ValueError:
                    return val_str
                    
            # Find the actual last content row
            max_r = sheet.max_row
            last_content_row = 1
            consecutive_empty = 0
            for r in range(1, max_r + 1):
                has_content = False
                for c in range(1, 35):
                    val = sheet.cell(row=r, column=c).value
                    if val is not None and str(val).strip() != "":
                        has_content = True
                        break
                if has_content:
                    last_content_row = r
                    consecutive_empty = 0
                else:
                    consecutive_empty += 1
                    if consecutive_empty >= 100:
                        break
                        
            next_row = last_content_row + 1
            for col_idx, val in enumerate(row_data, 1):
                sheet.cell(row=next_row, column=col_idx, value=cast_val(val))
                
            wb.save(CURRENT_DB_PATH)
            wb.close()
            
        else:
            return jsonify({'error': 'Unsupported file format for saving.'}), 400
            
        return jsonify({'success': True, 'message': 'Order saved successfully!'})
        
    except Exception as e:
        print(f"Error saving order: {e}")
        return jsonify({'error': f"Failed to save order: {str(e)}"}), 500


@app.route('/api/generate', methods=['POST'])
def generate_documents():
    data = request.json
    if not data:
        return jsonify({'error': 'No data provided'}), 400
        
    try:
        po_num = data.get('customer_po', 'UNKNOWN_PO').strip()
        rms_po = data.get('rms_po', 'UNKNOWN_RMS').strip()
        
        # Clean filenames (replace slashes and spaces)
        clean_po = "".join(c for c in po_num if c.isalnum() or c in '-_')
        clean_rms = "".join(c for c in rms_po if c.isalnum() or c in '-_')
        
        # Output filenames
        ps_filename = f"PS_{clean_po}_{clean_rms}.pdf"
        ci_filename = f"CI_{clean_po}_{clean_rms}.pdf"
        label_filename = f"Label_{clean_po}_{clean_rms}.docx"
        
        ps_path = os.path.join(OUTPUTS_DIR, ps_filename)
        ci_path = os.path.join(OUTPUTS_DIR, ci_filename)
        label_path = os.path.join(OUTPUTS_DIR, label_filename)
        
        # 1. Generate Printing Slip (PDF)
        generate_printing_slip(ps_path, data)
        
        # 2. Generate Commercial Invoice (PDF)
        generate_commercial_invoice(ci_path, data)
        
        # 3. Generate Part Labels (DOCX)
        # Parse qty from request data
        qty_str = data.get('qty', '4')
        try:
            qty = max(1, int(qty_str))
        except (ValueError, TypeError):
            qty = 4

        generate_docx_labels(
            template_path=TEMPLATE_DOCX_PATH,
            output_path=label_path,
            po_num=po_num,
            part_num=data.get('part_num', ''),
            description=data.get('part_desc', ''),
            qty=qty,
            items=data.get('items')
        )
        
        return jsonify({
            'success': True,
            'message': 'All documents generated successfully!',
            'files': {
                'printing_slip': {
                    'filename': ps_filename,
                    'path': ps_path
                },
                'commercial_invoice': {
                    'filename': ci_filename,
                    'path': ci_path
                },
                'part_labels': {
                    'filename': label_filename,
                    'path': label_path
                }
            }
        })
        
    except Exception as e:
        print(f"Error generating files: {e}")
        return jsonify({'error': str(e)}), 500



def ensure_receiving_headers_excel(sheet, header_row_idx):
    # Find existing headers
    max_col = sheet.max_column
    headers = [sheet.cell(row=header_row_idx, column=c).value for c in range(1, max_col + 1)]
    headers_clean = [str(h).strip().lower() if h is not None else "" for h in headers]
    
    boxes_idx = -1
    photo_idx = -1
    report_idx = -1
    
    for i, h in enumerate(headers_clean):
        if "boxes" in h or "no. of boxes" in h:
            boxes_idx = i + 1
        elif "photo" in h:
            photo_idx = i + 1
        elif "report" in h:
            report_idx = i + 1
            
    # If not found, append to the end
    curr_len = len(headers)
    if boxes_idx == -1:
        curr_len += 1
        sheet.cell(row=header_row_idx, column=curr_len, value="No. of boxes")
        boxes_idx = curr_len
    if photo_idx == -1:
        curr_len += 1
        sheet.cell(row=header_row_idx, column=curr_len, value="photo")
        photo_idx = curr_len
    if report_idx == -1:
        curr_len += 1
        sheet.cell(row=header_row_idx, column=curr_len, value="Report")
        report_idx = curr_len
        
    return boxes_idx, photo_idx, report_idx


def ensure_captures_header_excel(sheet, header_row_idx):
    max_col = sheet.max_column
    headers = [sheet.cell(row=header_row_idx, column=c).value for c in range(1, max_col + 1)]
    headers_clean = [str(h).strip().lower() if h is not None else "" for h in headers]
    
    captures_idx = -1
    for i, h in enumerate(headers_clean):
        if h == "captures":
            captures_idx = i + 1
            break
            
    if captures_idx == -1:
        new_col = len(headers) + 1
        sheet.cell(row=header_row_idx, column=new_col, value="Captures")
        captures_idx = new_col
        
    return captures_idx


def update_excel_records(file_path, sheet_name, header_row_idx, updates):
    wb = openpyxl.load_workbook(file_path)
    if not sheet_name or sheet_name not in wb.sheetnames:
        sheet = wb.active
    else:
        sheet = wb[sheet_name]
        
    boxes_idx, photo_idx, report_idx = ensure_receiving_headers_excel(sheet, header_row_idx)
    
    # Helper to clean/cast values for Excel
    def cast_val(val):
        val_str = str(val).strip()
        if val_str == "":
            return None
        try:
            if val_str.isdigit():
                return int(val_str)
            float_val = float(val_str)
            if float_val.is_integer():
                return int(float_val)
            return float_val
        except ValueError:
            return val_str
            
    for update in updates:
        row_id = int(update['row_id'])
        
        # Verify row_hash if provided
        row_hash = update.get('row_hash')
        if row_hash:
            current_rec = get_excel_row_record(sheet, row_id, boxes_idx, photo_idx, report_idx)
            current_hash = calculate_record_hash(current_rec)
            if current_hash != row_hash:
                raise Exception(f"Row {row_id} has been modified by another user. Please refresh page and re-apply changes.")
        
        if 'received_date' in update:
            sheet.cell(row=row_id, column=8, value=cast_val(update['received_date']))
        if 'inbound_carrier' in update:
            sheet.cell(row=row_id, column=9, value=cast_val(update['inbound_carrier']))
        if 'inbound_tracking' in update:
            sheet.cell(row=row_id, column=10, value=cast_val(update['inbound_tracking']))
            
        if 'inbound_l' in update:
            sheet.cell(row=row_id, column=11, value=cast_val(update['inbound_l']))
        if 'inbound_w' in update:
            sheet.cell(row=row_id, column=12, value=cast_val(update['inbound_w']))
        if 'inbound_h' in update:
            sheet.cell(row=row_id, column=13, value=cast_val(update['inbound_h']))
            
        if 'inbound_weight' in update:
            sheet.cell(row=row_id, column=14, value=cast_val(update['inbound_weight']))
            
        if boxes_idx != -1 and 'no_of_boxes' in update:
            sheet.cell(row=row_id, column=boxes_idx, value=cast_val(update['no_of_boxes']))
        if photo_idx != -1 and 'photo' in update:
            sheet.cell(row=row_id, column=photo_idx, value=cast_val(update['photo']))
        if report_idx != -1 and 'report' in update:
            sheet.cell(row=row_id, column=report_idx, value=cast_val(update['report']))
            
        # Outbound / shipping / order entry fields
        if 'part_received' in update:
            sheet.cell(row=row_id, column=3, value=cast_val(update['part_received']))
        if 'outbound_date' in update:
            sheet.cell(row=row_id, column=16, value=cast_val(update['outbound_date']))
        if 'customer_po' in update:
            sheet.cell(row=row_id, column=18, value=cast_val(update['customer_po']))
        if 'ship_to' in update:
            sheet.cell(row=row_id, column=20, value=cast_val(update['ship_to']))
        if 'line_num' in update:
            sheet.cell(row=row_id, column=21, value=cast_val(update['line_num']))
        if 'hs_code' in update:
            sheet.cell(row=row_id, column=22, value=cast_val(update['hs_code']))
        if 'outbound_l' in update:
            sheet.cell(row=row_id, column=25, value=cast_val(update['outbound_l']))
        if 'outbound_w' in update:
            sheet.cell(row=row_id, column=26, value=cast_val(update['outbound_w']))
        if 'outbound_h' in update:
            sheet.cell(row=row_id, column=27, value=cast_val(update['outbound_h']))
        if 'outbound_weight' in update:
            sheet.cell(row=row_id, column=28, value=cast_val(update['outbound_weight']))
        if 'shipped_date' in update:
            sheet.cell(row=row_id, column=23, value=cast_val(update['shipped_date']))
        if 'outbound_carrier' in update:
            sheet.cell(row=row_id, column=29, value=cast_val(update['outbound_carrier']))
        if 'outbound_tracking' in update:
            sheet.cell(row=row_id, column=30, value=cast_val(update['outbound_tracking']))
            
    wb.save(file_path)
    wb.close()


@app.route('/api/upload-photo', methods=['POST'])
def upload_photo():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    if file:
        filename = secure_filename(file.filename)
        import time
        timestamp = int(time.time())
        name, ext = os.path.splitext(filename)
        unique_filename = f"{name}_{timestamp}{ext}"
        
        uploads_dir = os.path.join(BASE_DIR, 'static', 'uploads')
        os.makedirs(uploads_dir, exist_ok=True)
        file_path = os.path.join(uploads_dir, unique_filename)
        file.save(file_path)
        
        return jsonify({
            'success': True,
            'filename': unique_filename,
            'url': f"/uploads/{unique_filename}"
        })


@app.route('/api/update-records', methods=['POST'])
def update_records():
    global CURRENT_DB_PATH, CURRENT_SHEET, CURRENT_HEADER_ROW
    data = request.json
    if not data or 'updates' not in data:
        return jsonify({'error': 'No updates provided'}), 400
        
    if not CURRENT_DB_PATH or not os.path.exists(CURRENT_DB_PATH):
        return jsonify({'error': 'No active database file loaded.'}), 400
        
    updates = data['updates']
    ext = os.path.splitext(CURRENT_DB_PATH)[1].lower()
    try:
        if ext in ['.xlsx', '.xls']:
            update_excel_records(CURRENT_DB_PATH, CURRENT_SHEET, CURRENT_HEADER_ROW, updates)
        else:
            return jsonify({'error': 'Unsupported file format.'}), 400
            
        return jsonify({'success': True, 'message': 'Records updated successfully!'})
    except Exception as e:
        print(f"Error updating records: {e}")
        return jsonify({'error': f"Failed to update records: {str(e)}"}), 500


@app.route('/api/send-email', methods=['POST'])
def send_email_route():
    data = request.json
    if not data or 'customer_email' not in data or 'record' not in data:
        return jsonify({'error': 'Missing customer_email or record data'}), 400
        
    customer_email = data['customer_email'].strip()
    record = data['record']
    
    if not customer_email:
        return jsonify({'error': 'Customer email cannot be empty'}), 400
        
    # Construct email details
    customer_po = record.get('customer_po', '').strip() or "N/A"
    rms_po = record.get('rms_po', '').strip() or "N/A"
    shipped_date = record.get('shipped_date', '').strip() or "N/A"
    order_date = record.get('outbound_date', '').strip() or "N/A"
    weight = record.get('outbound_weight', '').strip() or "N/A"
    out_l = record.get('outbound_l', '').strip() or "N/A"
    out_w = record.get('outbound_w', '').strip() or "N/A"
    out_h = record.get('outbound_h', '').strip() or "N/A"
    carrier = record.get('outbound_carrier', '').strip() or "N/A"
    tracking = record.get('outbound_tracking', '').strip() or "N/A"
    customer_name = record.get('customer', '').strip() or "Valued Customer"
    
    dimensions = f"{out_l} x {out_w} x {out_h} Inches"
    if out_l == "N/A" and out_w == "N/A" and out_h == "N/A":
        dimensions = "N/A"
    
    # Text fallback with only metadata (excluding Line Number)
    body = (
        f"Dear {customer_name},\n\n"
        f"This is a shipment notification for your order. Please find the shipping and order metadata below:\n\n"
        f"Customer P.O. / S.O. #: {customer_po}\n"
        f"Shipping Date: {shipped_date}\n"
        f"Order Date: {order_date}\n"
        f"Weight (Lbs.): {weight}\n"
        f"LxWxH (Inches): {dimensions}\n"
        f"Carrier: {carrier}\n"
        f"Tracking / Pro #: {tracking}\n\n"
        f"If you have any questions, please feel free to contact us.\n\n"
        f"Best regards,\n"
        f"RMS Team"
    )
    
    # Beautiful HTML table template (excluding Line Number)
    html_body = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<title>Shipment Notification</title>
</head>
<body style="font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif; color: #333333; background-color: #f4f6f8; padding: 20px 0; margin: 0; width: 100%; -webkit-font-smoothing: antialiased;">
  <div style="max-width: 600px; margin: 0 auto; background: #ffffff; border-radius: 8px; overflow: hidden; box-shadow: 0 4px 6px rgba(0, 0, 0, 0.05); border: 1px solid #e1e4e8;">
    <div style="background-color: #1a56db; color: #ffffff; padding: 24px; text-align: center;">
      <h2 style="margin: 0; font-size: 20px; font-weight: 600; letter-spacing: 0.5px;">Shipment Notification</h2>
    </div>
    <div style="padding: 30px;">
      <p style="font-size: 15px; line-height: 1.5; margin: 0 0 16px 0;">Dear {customer_name},</p>
      <p style="font-size: 15px; line-height: 1.5; margin: 0 0 24px 0;">This is a shipment notification for your order. Please find the shipping and order metadata below:</p>
      
      <table style="width: 100%; border-collapse: collapse; margin-top: 15px; margin-bottom: 25px;">
        <tr style="background-color: #ffffff;">
          <td style="padding: 12px 16px; text-align: left; border-bottom: 1px solid #eef2f6; font-weight: 600; color: #4b5563; width: 45%; font-size: 14px;">Customer P.O. / S.O. #</td>
          <td style="padding: 12px 16px; text-align: left; border-bottom: 1px solid #eef2f6; color: #1f2937; font-size: 14px;">{customer_po}</td>
        </tr>
        <tr style="background-color: #f9fafb;">
          <td style="padding: 12px 16px; text-align: left; border-bottom: 1px solid #eef2f6; font-weight: 600; color: #4b5563; width: 45%; font-size: 14px;">Shipping Date</td>
          <td style="padding: 12px 16px; text-align: left; border-bottom: 1px solid #eef2f6; color: #1f2937; font-size: 14px;">{shipped_date}</td>
        </tr>
        <tr style="background-color: #ffffff;">
          <td style="padding: 12px 16px; text-align: left; border-bottom: 1px solid #eef2f6; font-weight: 600; color: #4b5563; width: 45%; font-size: 14px;">Order Date</td>
          <td style="padding: 12px 16px; text-align: left; border-bottom: 1px solid #eef2f6; color: #1f2937; font-size: 14px;">{order_date}</td>
        </tr>
        <tr style="background-color: #f9fafb;">
          <td style="padding: 12px 16px; text-align: left; border-bottom: 1px solid #eef2f6; font-weight: 600; color: #4b5563; width: 45%; font-size: 14px;">Weight (Lbs.)</td>
          <td style="padding: 12px 16px; text-align: left; border-bottom: 1px solid #eef2f6; color: #1f2937; font-size: 14px;">{weight}</td>
        </tr>
        <tr style="background-color: #ffffff;">
          <td style="padding: 12px 16px; text-align: left; border-bottom: 1px solid #eef2f6; font-weight: 600; color: #4b5563; width: 45%; font-size: 14px;">LxWxH (Inches)</td>
          <td style="padding: 12px 16px; text-align: left; border-bottom: 1px solid #eef2f6; color: #1f2937; font-size: 14px;">{dimensions}</td>
        </tr>
        <tr style="background-color: #f9fafb;">
          <td style="padding: 12px 16px; text-align: left; border-bottom: 1px solid #eef2f6; font-weight: 600; color: #4b5563; width: 45%; font-size: 14px;">Carrier</td>
          <td style="padding: 12px 16px; text-align: left; border-bottom: 1px solid #eef2f6; color: #1f2937; font-size: 14px;">{carrier}</td>
        </tr>
        <tr style="background-color: #ffffff;">
          <td style="padding: 12px 16px; text-align: left; border-bottom: 1px solid #eef2f6; font-weight: 600; color: #4b5563; width: 45%; font-size: 14px;">Tracking / Pro #</td>
          <td style="padding: 12px 16px; text-align: left; border-bottom: 1px solid #eef2f6; color: #1f2937; font-size: 14px; font-weight: bold;">{tracking}</td>
        </tr>
      </table>
      
      <p style="font-size: 15px; line-height: 1.5; margin: 0 0 20px 0;">If you have any questions, please feel free to contact us.</p>
    </div>
    <div style="background-color: #f9fafb; padding: 20px; text-align: center; border-top: 1px solid #eef2f6; font-size: 13px; color: #6b7280;">
      Best regards,<br>
      <span style="font-weight: bold; color: #1f2937;">RMS Team</span>
    </div>
  </div>
</body>
</html>"""

    api_key = "a17cfc3e0583860e02fa41d4d666630a-9889a0ac-8a080308"
    domain = "mail.devakash.in"
    url = f"https://api.mailgun.net/v3/{domain}/messages"
    
    import base64
    import urllib.request
    import urllib.parse
    
    auth_str = f"api:{api_key}"
    auth_b64 = base64.b64encode(auth_str.encode('utf-8')).decode('utf-8')
    headers = {
        "Authorization": f"Basic {auth_b64}",
        "Content-Type": "application/x-www-form-urlencoded"
    }
    
    # 1. Send to Customer
    customer_subject = f"Shipment Notification - RMS P.O. #{rms_po} / Customer P.O. #{customer_po}"
    post_data_customer = {
        "from": f"Mailgun Sandbox <postmaster@{domain}>",
        "to": customer_email,
        "subject": customer_subject,
        "text": body,
        "html": html_body
    }
    encoded_customer = urllib.parse.urlencode(post_data_customer).encode('utf-8')
    req_customer = urllib.request.Request(url, data=encoded_customer, headers=headers, method="POST")
    
    # 2. Send to Team
    team_subject = f"Copy of shipment email sent Customer: {customer_po}"
    team_emails = ["rajan@rmsint.net", "sales@rmsint.net"]
    team_to_str = ", ".join(team_emails)
    post_data_team = {
        "from": f"Mailgun Sandbox <postmaster@{domain}>",
        "to": team_to_str,
        "subject": team_subject,
        "text": body,
        "html": html_body
    }
    encoded_team = urllib.parse.urlencode(post_data_team).encode('utf-8')
    req_team = urllib.request.Request(url, data=encoded_team, headers=headers, method="POST")
    
    try:
        # Send customer email
        with urllib.request.urlopen(req_customer) as response_cust:
            res_body_cust = response_cust.read().decode('utf-8')
            
        # Send team email (fail-safe wrapper)
        try:
            with urllib.request.urlopen(req_team) as response_team:
                pass
        except Exception as e_team:
            print(f"Error sending email to team: {e_team}")
            
        return jsonify({
            'success': True,
            'message': 'Email sent successfully!',
            'response': res_body_cust
        })
    except Exception as e:
        print(f"Error sending email: {e}")
        error_body = ""
        if hasattr(e, 'read'):
            try:
                error_body = e.read().decode('utf-8')
            except Exception:
                pass
        err_msg = f"{str(e)}: {error_body}" if error_body else str(e)
        return jsonify({'error': f"Failed to send email: {err_msg}"}), 500


def load_dotenv():
    env_path = os.path.join(BASE_DIR, '.env')
    if os.path.exists(env_path):
        try:
            with open(env_path, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith('#'):
                        continue
                    if '=' in line:
                        key, val = line.split('=', 1)
                        os.environ[key.strip()] = val.strip()
        except Exception as e:
            print(f"Error loading env file: {e}")

load_dotenv()


def fetch_from_scanner_api(endpoint, method='GET', payload=None):
    import urllib.request
    import json
    import os
    
    scanner_url = os.environ.get('SCANNER_APP_URL', 'https://rms-scanner.vercel.app')
    token = os.environ.get('SHARED_API_TOKEN', 'd4b8e21a-7b3e-4d56-bc98-fa39e6a39281')
    
    url = f"{scanner_url.rstrip('/')}{endpoint}"
    req = urllib.request.Request(url, method=method)
    req.add_header('Authorization', f'Bearer {token}')
    req.add_header('Content-Type', 'application/json')
    
    data_bytes = None
    if payload is not None:
        data_bytes = json.dumps(payload).encode('utf-8')
        
    try:
        import ssl
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        
        with urllib.request.urlopen(req, data=data_bytes, timeout=10, context=ctx) as response:
            res_data = response.read().decode('utf-8')
            return json.loads(res_data)
    except Exception as e:
        print(f"Error fetching from scanner API {url}: {e}")
        return {'success': False, 'error': str(e)}


@app.route('/api/check-scans')
def check_scans():
    res = fetch_from_scanner_api('/api/scan')
    if not res.get('success'):
        return jsonify({'error': f"Failed to fetch scans from scanner app: {res.get('error')}"}), 500
    return jsonify({'success': True, 'scans': res.get('scans', [])})


@app.route('/api/shipping-captures')
def get_shipping_captures():
    res = fetch_from_scanner_api('/api/shipping-capture')
    if not res.get('success'):
        return jsonify({'error': f"Failed to fetch captures from scanner app: {res.get('error')}"}), 500
    return jsonify({'success': True, 'captures': res.get('captures', [])})


@app.route('/api/link-shipping-images', methods=['POST'])
def link_shipping_images():
    import base64
    import os
    
    data = request.json
    if not data or 'image_ids' not in data or 'folder_path' not in data:
        return jsonify({'error': 'Missing required parameters'}), 400
        
    image_ids = data['image_ids']
    folder_path = data['folder_path'].strip()
    row_id = data.get('row_id')
    customer_po = data.get('customer_po', '').strip() or "capture"
    
    if not image_ids:
        return jsonify({'error': 'No images selected'}), 400
    if not folder_path:
        return jsonify({'error': 'Folder path cannot be empty'}), 400
        
    try:
        try:
            os.makedirs(folder_path, exist_ok=True)
        except Exception as e_fs:
            return jsonify({'error': f"Failed to access/create directory {folder_path}: {str(e_fs)}"}), 500
            
        # Fetch all captures via Next.js API
        res = fetch_from_scanner_api('/api/shipping-capture')
        if not res.get('success'):
            return jsonify({'error': f"Failed to fetch captures: {res.get('error')}"}), 500
            
        captures = {c['id']: c for c in res.get('captures', [])}
        
        import time
        saved_count = 0
        saved_file_paths = []
        linked_ids = []
        
        for idx, img_id in enumerate(image_ids):
            doc = captures.get(img_id)
            if not doc:
                continue
                
            image_data_url = doc.get('image', '')
            if not image_data_url or ',' not in image_data_url:
                continue
                
            # Decode base64
            header, base64_data = image_data_url.split(',', 1)
            image_bytes = base64.b64decode(base64_data)
            
            # Save locally
            timestamp = int(time.time())
            filename = f"capture_{timestamp}_{idx}_{img_id}.jpg"
            file_path = os.path.join(folder_path, filename)
            
            with open(file_path, 'wb') as f:
                f.write(image_bytes)
            saved_count += 1
            saved_file_paths.append(file_path)
            linked_ids.append(img_id)
            
        # Call Next.js API to delete these linked images from MongoDB
        if linked_ids:
            del_res = fetch_from_scanner_api('/api/shipping-capture', method='POST', payload={
                'action': 'delete',
                'image_ids': linked_ids
            })
            if not del_res.get('success'):
                print(f"Warning: Failed to delete linked captures from scanner app: {del_res.get('error')}")
                
        # Construct and write Excel hyperlink formulas
        if saved_file_paths and row_id and os.path.exists(CURRENT_DB_PATH):
            formulas = []
            for i, fp in enumerate(saved_file_paths):
                win_path = os.path.normpath(fp).replace('/', '\\')
                file_url = f"file:///{win_path}"
                friendly_name = f"{customer_po}.jpg" if len(saved_file_paths) == 1 else f"{customer_po}_{i + 1}.jpg"
                formulas.append(f'HYPERLINK("{file_url}", "{friendly_name}")')
                
            formula_val = "=" + " & \", \" & ".join(formulas)
            
            wb = openpyxl.load_workbook(CURRENT_DB_PATH)
            sheet = wb[CURRENT_SHEET] if CURRENT_SHEET in wb.sheetnames else wb.active
            
            captures_col = ensure_captures_header_excel(sheet, CURRENT_HEADER_ROW)
            sheet.cell(row=int(row_id), column=captures_col, value=formula_val)
            
            wb.save(CURRENT_DB_PATH)
            wb.close()
            
        return jsonify({
            'success': True,
            'message': f"Successfully linked {saved_count} images to shipping. Saved in {folder_path}."
        })
    except Exception as e:
        print(f"Error linking shipping images: {e}")
        return jsonify({'error': f"Failed to link images: {str(e)}"}), 500


@app.route('/api/browse-directory', methods=['POST'])
def browse_directory():
    try:
        import tkinter as tk
        from tkinter import filedialog
        
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        
        directory = filedialog.askdirectory(title="Select Shipping Captures Base Directory")
        root.destroy()
        
        if directory:
            return jsonify({'success': True, 'directory': directory})
        else:
            return jsonify({'success': False, 'message': 'No directory selected'})
    except Exception as e:
        print(f"Error opening directory dialog: {e}")
        return jsonify({'error': f"Could not open directory browser: {str(e)}"}), 500


@app.route('/api/generated-files')
def get_generated_files():
    try:
        files = []
        if os.path.exists(OUTPUTS_DIR):
            for filename in os.listdir(OUTPUTS_DIR):
                if filename.startswith('.'):
                    continue
                path = os.path.join(OUTPUTS_DIR, filename)
                if os.path.isfile(path):
                    stat = os.stat(path)
                    files.append({
                        'name': filename,
                        'path': path,
                        'size': stat.st_size,
                        'mtime': stat.st_mtime
                    })
            # Sort by modification time, newest first
            files.sort(key=lambda x: x['mtime'], reverse=True)
        return jsonify(files)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/open-file')
def open_file():
    filepath = request.args.get('path')
    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': 'File not found'}), 404
        
    try:
        # Open file using OS default application
        if sys.platform == 'darwin':
            os.system(f'open "{filepath}"')
        elif sys.platform == 'win32':
            os.startfile(filepath)
        else:
            os.system(f'xdg-open "{filepath}"')
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/open-outputs-folder')
def open_outputs_folder():
    try:
        if sys.platform == 'darwin':
            os.system(f'open "{OUTPUTS_DIR}"')
        elif sys.platform == 'win32':
            os.startfile(OUTPUTS_DIR)
        else:
            os.system(f'xdg-open "{OUTPUTS_DIR}"')
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def open_browser():
    webbrowser.open_new('http://localhost:5001')


if __name__ == '__main__':
    # Delay opening browser for 1 second to ensure Flask starts first
    Timer(1.0, open_browser).start()
    app.run(host='127.0.0.1', port=5001, debug=False)
