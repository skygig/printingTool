import os
import json
import unittest
import unittest.mock
from app import app

class TestDocumentGeneratorApp(unittest.TestCase):
    def setUp(self):
        app.config['TESTING'] = True
        self.app = app.test_client()



    def test_api_records_endpoint(self):
        print("Testing /api/records endpoint...")
        response = self.app.get('/api/records')
        self.assertEqual(response.status_code, 200)
        data = json.loads(response.data)
        self.assertTrue(isinstance(data, list))
        self.assertTrue(len(data) > 0)
        print(f"  API returned {len(data)} records successfully.")

    def test_api_generate_endpoint(self):
        print("Testing /api/generate endpoint...")
        test_payload = {
            'date': '06/07/2026',
            'order_date': '05/01/2026',
            'customer_po': '42300191168',
            'rms_po': '9679',
            'ref_num': '191168-9679',
            'tax_id': '36-4426459',
            'line_num': '1',
            'part_num': '1-503-24-065',
            'part_desc': 'TEFLON SEAL (TEST_RUN)',
            'description': 'TEFLON SEAL (TEST_RUN)',
            'qty': '6',
            'backordered': '0',
            'hs_code': '3926.90.9985',
            'amount': '250.00',
            'weight': '2LBS',
            'size': '13x11x5',
            'ship_to_address': 'GE Energy Parts\nEnergy Services - Energy Parts COE\n4955 Mason Road\nAtlanta, GA 30349',
            'sold_to_address': 'GE Energy Parts\nEnergy Services - Energy Parts COE\n4955 Mason Road\nAtlanta, GA 30349',
            'notes': ['1. Test Note 1', '2. Test Note 2'],
            'free_replacement_note': 'Free replacement'
        }
        
        response = self.app.post(
            '/api/generate',
            data=json.dumps(test_payload),
            content_type='application/json'
        )
        self.assertEqual(response.status_code, 200)
        data = json.loads(response.data)
        self.assertTrue(data['success'])
        self.assertIn('files', data)
        
        # Verify files are written to disk
        ps_path = data['files']['printing_slip']['path']
        ci_path = data['files']['commercial_invoice']['path']
        label_path = data['files']['part_labels']['path']
        
        self.assertTrue(os.path.exists(ps_path), f"Packing slip file not found: {ps_path}")
        self.assertTrue(os.path.exists(ci_path), f"Commercial Invoice file not found: {ci_path}")
        self.assertTrue(os.path.exists(label_path), f"Label file not found: {label_path}")
        
        print("  All three files generated and written to Outputs successfully.")
        print(f"    PS: {ps_path}")
        print(f"    CI: {ci_path}")
        print(f"    Label: {label_path}")

    def test_api_generate_endpoint_multiple_items(self):
        print("Testing /api/generate endpoint with multiple items...")
        test_payload = {
            'date': '06/07/2026',
            'order_date': '05/01/2026',
            'customer_po': '42300191168-MULTI',
            'rms_po': '9679',
            'ref_num': '191168-9679-MULTI',
            'tax_id': '36-4426459',
            'weight': '4LBS',
            'size': '13x11x10',
            'ship_to_address': 'GE Energy Parts\nAtlanta, GA 30349',
            'sold_to_address': 'GE Energy Parts\nAtlanta, GA 30349',
            'notes': ['1. Multi Item test note'],
            'free_replacement_note': 'Free replacement',
            'items': [
                {
                    'rms_po': '9679',
                    'line_num': '1',
                    'part_received': 'QTY 2 PN 1-503-24-065 TEFLON SEAL',
                    'part_num': '1-503-24-065',
                    'part_desc': 'TEFLON SEAL',
                    'qty': '2',
                    'backordered': '0',
                    'hs_code': '3926.90.9985',
                    'amount': '120.00'
                },
                {
                    'rms_po': '9680',
                    'line_num': '2',
                    'part_received': 'QTY 3 PN 2-300-11-233 GASKET',
                    'part_num': '2-300-11-233',
                    'part_desc': 'GASKET',
                    'qty': '3',
                    'backordered': '1',
                    'hs_code': '3926.90.9985',
                    'amount': '80.00'
                }
            ]
        }
        
        response = self.app.post(
            '/api/generate',
            data=json.dumps(test_payload),
            content_type='application/json'
        )
        self.assertEqual(response.status_code, 200)
        data = json.loads(response.data)
        self.assertTrue(data['success'])
        self.assertIn('files', data)
        
        # Verify files are written to disk
        ps_path = data['files']['printing_slip']['path']
        ci_path = data['files']['commercial_invoice']['path']
        label_path = data['files']['part_labels']['path']
        
        self.assertTrue(os.path.exists(ps_path), f"Packing slip file not found: {ps_path}")
        self.assertTrue(os.path.exists(ci_path), f"Commercial Invoice file not found: {ci_path}")
        self.assertTrue(os.path.exists(label_path), f"Label file not found: {label_path}")
        
        print("  All multiple items files generated and written to Outputs successfully.")

    def test_database_picker_api(self):
        print("Testing Database Picker and Upload endpoints...")
        
        # Test 1: Get DB Status
        response = self.app.get('/api/db-status')
        self.assertEqual(response.status_code, 200)
        status = json.loads(response.data)
        self.assertIn('filename', status)
        self.assertIn('sheet_name', status)
        self.assertIn('sheets', status)
        print("  /api/db-status verified.")
        
        # Test 2: Verify CSV upload is rejected
        import io
        csv_content = "Inbound Date,RMS P.O.\n06/07/2026,9999\n"
        data = {
            'file': (io.BytesIO(csv_content.encode('utf-8')), 'test_mock_db.csv')
        }
        response = self.app.post(
            '/api/upload-database',
            data=data,
            content_type='multipart/form-data'
        )
        self.assertEqual(response.status_code, 400)
        res = json.loads(response.data)
        self.assertIn('Unsupported file format', res['error'])
        print("  /api/upload-database CSV rejection verified.")

        # Test 3: Upload a sample XLSX database
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MainSheet"
        headers = [
            "Inbound Date", "RMS P.O.", "Part Received/QTY/PN", "Vendor", "Promise Date",
            "Inbound Notes", "Vendor Contact", "Received Date", "Inbound Carrier", "Inbound Tracking",
            "Inbound L", "Inbound W", "Inbound H", "Inbound Weight", "Inbound Charges",
            "Outbound Date", "Customer", "Customer P.O.", "RMS Invoice", "Ship To",
            "Line Num", "HS Code", "Shipped Date", "Invoice Status", "Outbound L",
            "Outbound W", "Outbound H", "Outbound Weight", "Outbound Carrier", "Outbound Tracking",
            "Crating Charges", "Shipping Charges", "Customer Contact", "Outbound Notes"
        ]
        ws.append(headers)
        row_data = [""] * 34
        row_data[1] = "8888"
        row_data[2] = "PART_B QTY 5 PN 789-012"
        row_data[16] = "CUST_B"
        row_data[17] = "42300000001"
        ws.append(row_data)
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        data_xlsx = {
            'file': (excel_file, 'test_mock_db.xlsx')
        }
        
        response = self.app.post(
            '/api/upload-database',
            data=data_xlsx,
            content_type='multipart/form-data'
        )
        self.assertEqual(response.status_code, 200)
        upload_res = json.loads(response.data)
        self.assertTrue(upload_res['success'])
        self.assertEqual(upload_res['type'], 'excel')
        self.assertEqual(upload_res['filename'], 'test_mock_db.xlsx')
        self.assertIn('sheets', upload_res)
        self.assertEqual(upload_res['sheets'], ['MainSheet'])
        self.assertTrue(len(upload_res['records']) > 0)
        self.assertEqual(upload_res['records'][0]['rms_po'], '8888')
        print("  /api/upload-database XLSX upload verified.")
        
        # Cleanup mock file
        mock_file_path_xlsx = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'test_mock_db.xlsx')
        if os.path.exists(mock_file_path_xlsx):
            os.remove(mock_file_path_xlsx)

        # Test 4: Load a sheet with a custom header row number
        wb_offset = openpyxl.Workbook()
        ws_offset = wb_offset.active
        ws_offset.title = "OffsetSheet"
        ws_offset.append(["Garbage Title Row", "", "", ""]) # Row 1
        ws_offset.append(headers) # Row 2
        row_offset = [""] * 34
        row_offset[1] = "7777"
        row_offset[2] = "PART_C QTY 3 PN 345-678"
        row_offset[16] = "CUST_C"
        row_offset[17] = "42300000002"
        ws_offset.append(row_offset) # Row 3
        
        excel_file_offset = io.BytesIO()
        wb_offset.save(excel_file_offset)
        excel_file_offset.seek(0)
        
        data_offset = {
            'file': (excel_file_offset, 'test_header_offset.xlsx')
        }
        
        response = self.app.post(
            '/api/upload-database',
            data=data_offset,
            content_type='multipart/form-data'
        )
        self.assertEqual(response.status_code, 200)
        
        response = self.app.post(
            '/api/load-sheet',
            data=json.dumps({
                'sheet_name': 'OffsetSheet',
                'header_row': 2
            }),
            content_type='application/json'
        )
        self.assertEqual(response.status_code, 200)
        load_res = json.loads(response.data)
        self.assertTrue(load_res['success'])
        self.assertEqual(load_res['header_row'], 2)
        self.assertTrue(len(load_res['records']) > 0)
        self.assertEqual(load_res['records'][0]['row_id'], 3)
        self.assertEqual(load_res['records'][0]['rms_po'], '7777')
        print("  /api/load-sheet with header_row offset verified.")
        
        # Test 5: Verify db_config.json is created and loaded correctly
        config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'db_config.json')
        self.assertTrue(os.path.exists(config_path), "db_config.json was not created.")
        
        with open(config_path, 'r', encoding='utf-8') as f:
            config_data = json.load(f)
            
        self.assertEqual(config_data['filename'], 'test_header_offset.xlsx')
        self.assertEqual(config_data['sheet_name'], 'OffsetSheet')
        self.assertEqual(config_data['header_row'], 2)
        print("  db_config.json creation and correct contents verified.")
        
        # Verify startup loader can parse this configuration
        from app import load_db_config
        import app as app_module
        app_module.CURRENT_DB_NAME = None
        app_module.CURRENT_SHEET = None
        app_module.CURRENT_HEADER_ROW = 1
        
        # Load config
        load_db_config()
        self.assertEqual(app_module.CURRENT_DB_NAME, 'test_header_offset.xlsx')
        self.assertEqual(app_module.CURRENT_SHEET, 'OffsetSheet')
        self.assertEqual(app_module.CURRENT_HEADER_ROW, 2)
        print("  app.load_db_config() startup reload verified.")
        
        # Cleanup mock offset file and config file
        mock_file_path_offset = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'test_header_offset.xlsx')
        if os.path.exists(mock_file_path_offset):
            os.remove(mock_file_path_offset)
        if os.path.exists(config_path):
            os.remove(config_path)

    def test_save_order_endpoint(self):
        print("Testing /api/save-order endpoint...")
        # Save order to a mock Excel sheet
        import openpyxl
        
        headers = [
            "Date", "RMS P.O", "Part Received", "Vendor", "Promise date", "Notes", "Vendor Contact", 
            "Recieved Date", "Carrier", "Trackin/Pro#", "L", "W", "H", "Weight", "Inbound Shipping Charges", 
            "Date", "Customer", "Customer P.O", "RMS Invoice #", "Ship To", "Line #", "HS code", 
            "Shipped", "Invoice Status", "L", "W", "H", "Weight", "Carrier", "Tracking/ Pro#", 
            "Crating/ Handling charges", "Ship out Charges", "Customer contact", "Notes"
        ]
        
        # Set app global active DB
        import app as app_module
        old_db_path = app_module.CURRENT_DB_PATH
        old_db_name = app_module.CURRENT_DB_NAME
        
        base_dir = os.path.dirname(os.path.abspath(__file__))
        mock_xlsx_path = os.path.join(base_dir, "test_mock_save.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MainSheet"
        ws.append(headers)
        wb.save(mock_xlsx_path)
        wb.close()
        
        app_module.CURRENT_DB_PATH = mock_xlsx_path
        app_module.CURRENT_DB_NAME = "test_mock_save.xlsx"
        app_module.CURRENT_SHEET = "MainSheet"
        app_module.CURRENT_HEADER_ROW = 1
        
        # Call save-order endpoint
        test_payload = {
            'inbound_date': '06/07/2026',
            'rms_po': '1111',
            'part_received': 'QTY 2 PN PN-TEST PART',
            'vendor': 'VEND_TEST',
            'customer': 'CUST_TEST',
            'customer_po': 'PO-TEST',
            'invoice_status': ''
        }
        
        response = self.app.post(
            '/api/save-order',
            data=json.dumps(test_payload),
            content_type='application/json'
        )
        self.assertEqual(response.status_code, 200)
        res_data = json.loads(response.data)
        self.assertTrue(res_data['success'])
        
        # Verify Excel sheet records
        records_xlsx = app_module.parse_excel_database(mock_xlsx_path, "MainSheet", 1)
        self.assertEqual(len(records_xlsx), 1)
        self.assertEqual(records_xlsx[0]['rms_po'], '1111')
        self.assertEqual(records_xlsx[0]['customer_po'], 'PO-TEST')
        
        # Restore old globals
        app_module.CURRENT_DB_PATH = old_db_path
        app_module.CURRENT_DB_NAME = old_db_name
        app_module.CURRENT_SHEET = None
        app_module.CURRENT_HEADER_ROW = 1
        
        # Clean up files
        if os.path.exists(mock_xlsx_path):
            os.remove(mock_xlsx_path)
            
        print("  /api/save-order endpoints tests verified successfully for XLSX!")

    def test_optimistic_locking(self):
        print("Testing Optimistic Concurrency Control...")
        import openpyxl
        base_dir = os.path.dirname(os.path.abspath(__file__))
        mock_xlsx_path = os.path.join(base_dir, "test_mock_lock.xlsx")
        
        headers = [
            "Date", "RMS P.O", "Part Received", "Vendor", "Promise date", "Notes", "Vendor Contact", 
            "Recieved Date", "Carrier", "Trackin/Pro#", "L", "W", "H", "Weight", "Inbound Shipping Charges", 
            "Date", "Customer", "Customer P.O", "RMS Invoice #", "Ship To", "Line #", "HS code", 
            "Shipped", "Invoice Status", "L", "W", "H", "Weight", "Carrier", "Tracking/ Pro#", 
            "Crating/ Handling charges", "Ship out Charges", "Customer contact", "Notes"
        ]
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MainSheet"
        ws.append(headers)
        row_data = [""] * 34
        row_data[1] = "5555"
        row_data[2] = "PN-TEST"
        ws.append(row_data)
        wb.save(mock_xlsx_path)
        wb.close()
        
        import app as app_module
        old_db_path = app_module.CURRENT_DB_PATH
        old_db_name = app_module.CURRENT_DB_NAME
        app_module.CURRENT_DB_PATH = mock_xlsx_path
        app_module.CURRENT_DB_NAME = "test_mock_lock.xlsx"
        app_module.CURRENT_SHEET = "MainSheet"
        app_module.CURRENT_HEADER_ROW = 1
        
        # Parse it to get current records and their row hash
        records = app_module.parse_excel_database(mock_xlsx_path, "MainSheet", 1)
        self.assertEqual(len(records), 1)
        rec = records[0]
        correct_hash = rec['row_hash']
        
        # Try updating with wrong hash
        payload_wrong = {
            'updates': [{
                'row_id': rec['row_id'],
                'row_hash': 'wrong_hash_value',
                'customer_po': 'NEW-PO'
            }]
        }
        response = self.app.post(
            '/api/update-records',
            data=json.dumps(payload_wrong),
            content_type='application/json'
        )
        self.assertEqual(response.status_code, 500)
        res_data = json.loads(response.data)
        self.assertIn('modified by another user', res_data['error'])
        print("  Blocked update with mismatching row hash verified.")
        
        # Try updating with correct hash
        payload_correct = {
            'updates': [{
                'row_id': rec['row_id'],
                'row_hash': correct_hash,
                'customer_po': 'NEW-PO',
                'outbound_tracking': 'TRACKING-12345'
            }]
        }
        response = self.app.post(
            '/api/update-records',
            data=json.dumps(payload_correct),
            content_type='application/json'
        )
        self.assertEqual(response.status_code, 200)
        res_data = json.loads(response.data)
        self.assertTrue(res_data['success'])
        
        # Verify value in Excel
        import openpyxl
        wb_check = openpyxl.load_workbook(mock_xlsx_path)
        ws_check = wb_check["MainSheet"]
        saved_tracking = ws_check.cell(row=rec['row_id'], column=30).value
        self.assertEqual(saved_tracking, 'TRACKING-12345')
        wb_check.close()
        
        print("  Allowed update with matching row hash verified.")
        
        # Restore old globals and cleanup
        app_module.CURRENT_DB_PATH = old_db_path
        app_module.CURRENT_DB_NAME = old_db_name
        app_module.CURRENT_SHEET = None
        app_module.CURRENT_HEADER_ROW = 1
        if os.path.exists(mock_xlsx_path):
            os.remove(mock_xlsx_path)

    @unittest.mock.patch('urllib.request.urlopen')
    def test_send_email_route(self, mock_urlopen):
        print("Testing /api/send-email endpoint...")
        # Mock successful Mailgun API response
        mock_response = unittest.mock.MagicMock()
        mock_response.__enter__.return_value = mock_response
        mock_response.getcode.return_value = 200
        mock_response.read.return_value = b'{"id":"<test-id>","message":"Queued. Thank you."}'
        mock_urlopen.return_value = mock_response

        payload = {
            'customer_email': 'test@customer.com',
            'record': {
                'customer_po': '12345',
                'rms_po': '9999',
                'line_num': '1',
                'shipped_date': '05/20/2026',
                'outbound_date': '04/14/2026',
                'outbound_weight': '2 LBS',
                'outbound_l': '13',
                'outbound_w': '11',
                'outbound_h': '5',
                'outbound_carrier': 'UPS',
                'outbound_tracking': '1Z9999',
                'customer': 'Test Customer',
                'part_received': 'QTY 5 PN 12345 Test Item'
            }
        }
        response = self.app.post(
            '/api/send-email',
            data=json.dumps(payload),
            content_type='application/json'
        )
        self.assertEqual(response.status_code, 200)
        res_data = json.loads(response.data)
        self.assertTrue(res_data['success'])
        self.assertIn('Queued. Thank you.', res_data['response'])
        print("  /api/send-email success flow verified.")

        # Test error handling when Mailgun fails
        mock_urlopen.side_effect = Exception("Mailgun connection error")
        response = self.app.post(
            '/api/send-email',
            data=json.dumps(payload),
            content_type='application/json'
        )
        self.assertEqual(response.status_code, 500)
        res_data = json.loads(response.data)
        self.assertIn('Failed to send email', res_data['error'])
        print("  /api/send-email failure flow verified.")

    def test_authentication_flow(self):
        print("Testing Authentication and Authorization Flow...")
        
        # Turn off testing mode configuration so the middleware runs
        app.config['TESTING'] = False
        
        # 1. Access protected route without logging in -> 401
        response = self.app.get('/api/records')
        self.assertEqual(response.status_code, 401)
        
        # 2. Login with invalid credentials -> 401
        response = self.app.post(
            '/api/login',
            data=json.dumps({'username': 'raj', 'password': 'wrongpassword'}),
            content_type='application/json'
        )
        self.assertEqual(response.status_code, 401)
        
        # 3. Login as employee (accounting) -> 200
        response = self.app.post(
            '/api/login',
            data=json.dumps({'username': 'accounting', 'password': 'Acc12361$'}),
            content_type='application/json'
        )
        self.assertEqual(response.status_code, 200)
        data = json.loads(response.data)
        self.assertTrue(data['success'])
        self.assertEqual(data['role'], 'employee')
        
        # 4. Access shipping records as employee -> 200
        response = self.app.get('/api/records')
        self.assertEqual(response.status_code, 200)
        
        # 5. Access save-order (restricted to admin) as employee -> 403 Forbidden
        response = self.app.post(
            '/api/save-order',
            data=json.dumps({}),
            content_type='application/json'
        )
        self.assertEqual(response.status_code, 403)
        
        # 6. Logout -> 200
        response = self.app.post('/api/logout')
        self.assertEqual(response.status_code, 200)
        
        # 7. Check auth-status -> False
        response = self.app.get('/api/auth-status')
        self.assertEqual(response.status_code, 200)
        status = json.loads(response.data)
        self.assertFalse(status['authenticated'])
        
        # 8. Login as admin (raj) -> 200
        response = self.app.post(
            '/api/login',
            data=json.dumps({'username': 'raj', 'password': 'Plainfield1$'}),
            content_type='application/json'
        )
        self.assertEqual(response.status_code, 200)
        
        # 9. Access save-order as admin. Bypasses 403 Forbidden and hits route logic
        # (which returns 400 since there is no active DB path loaded in app context for this test)
        response = self.app.post(
            '/api/save-order',
            data=json.dumps({}),
            content_type='application/json'
        )
        self.assertNotEqual(response.status_code, 403)
        
        # Restore testing configuration
        app.config['TESTING'] = True
        print("  Authentication and role-based authorization tests passed successfully!")

    @unittest.mock.patch('pymongo.MongoClient')
    def test_shipping_captures_flow(self, mock_client_class):
        print("Testing /api/shipping-captures and /api/link-shipping-images endpoints...")
        
        # Mock mongodb find
        mock_client = unittest.mock.MagicMock()
        mock_client_class.return_value = mock_client
        mock_db = mock_client.__getitem__.return_value
        mock_collection = mock_db.__getitem__.return_value
        
        from datetime import datetime
        mock_doc = {
            '_id': '64b0f9f3f9f3f9f3f9f3f9f3',
            'image': 'data:image/jpeg;base64,dGVzdA==', # base64 of 'test'
            'capturedAt': datetime(2026, 7, 16, 3, 20, 28),
            'username': 'raj@rmsint.net'
        }
        
        # Configure find cursor
        mock_cursor = [mock_doc]
        mock_sorted = unittest.mock.MagicMock()
        mock_sorted.__iter__.return_value = iter(mock_cursor)
        mock_collection.find.return_value.sort.return_value = mock_sorted
        
        # 1. Test GET /api/shipping-captures
        response = self.app.get('/api/shipping-captures')
        self.assertEqual(response.status_code, 200)
        res_data = json.loads(response.data)
        self.assertTrue(res_data['success'])
        self.assertEqual(len(res_data['captures']), 1)
        self.assertEqual(res_data['captures'][0]['id'], '64b0f9f3f9f3f9f3f9f3f9f3')
        print("  /api/shipping-captures GET verified.")
        
        # 2. Test POST /api/link-shipping-images
        mock_collection.find_one.return_value = mock_doc
        
        # Mock os.makedirs and open to prevent actual file writes
        with unittest.mock.patch('os.makedirs') as mock_makedirs, \
             unittest.mock.patch('builtins.open', unittest.mock.mock_open()) as mock_file:
            
            payload = {
                'image_ids': ['64b0f9f3f9f3f9f3f9f3f9f3'],
                'folder_path': 'Z:/shipping_captures/Test Customer_PO12345',
                'row_id': 2,
                'customer_po': 'PO12345'
            }
            
            response = self.app.post(
                '/api/link-shipping-images',
                data=json.dumps(payload),
                content_type='application/json'
            )
            
            self.assertEqual(response.status_code, 200)
            res_data = json.loads(response.data)
            self.assertTrue(res_data['success'])
            self.assertIn('Successfully linked 1 images', res_data['message'])
            
            mock_makedirs.assert_called_once_with('Z:/shipping_captures/Test Customer_PO12345', exist_ok=True)
            mock_file.assert_called_once()
            mock_collection.delete_one.assert_called_once()
            print("  /api/link-shipping-images POST verified.")

if __name__ == '__main__':
    unittest.main()

